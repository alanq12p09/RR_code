"""
邮件发送模块：按 GEO 拆分推荐结果，每个 GEO 只收到自己的数据。
支持 Outlook COM 和 SMTP 两种方式，通过 config.MAIL_METHOD 切换。
"""
import os
import smtplib
import logging
import tempfile
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)


# ====================================================================
# 方式一：Outlook COM
# ====================================================================

def send_via_outlook(mail_to, mail_cc, subject, body, attachment_path=None):
    """通过本机 Outlook 客户端发送邮件。"""
    try:
        import win32com.client
    except ImportError:
        raise ImportError(
            "使用 Outlook 方式发邮件需要 pywin32 库，请安装: pip install pywin32\n"
            "或者在 config.py 中将 MAIL_METHOD 改为 'smtp' 使用 SMTP 方式。"
        )

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)

    mail.To = "; ".join(mail_to)
    if mail_cc:
        mail.CC = "; ".join(mail_cc)
    mail.Subject = subject
    mail.Body = body

    if attachment_path:
        file_path = Path(attachment_path)
        if not file_path.exists():
            raise FileNotFoundError(f"附件不存在: {attachment_path}")
        mail.Attachments.Add(str(file_path.resolve()))
        logger.info(f"已添加附件: {file_path.name} ({file_path.stat().st_size / 1024:.1f} KB)")

    mail.Send()
    logger.info(f"邮件已通过 Outlook 发送 -> {mail_to}")


# ====================================================================
# 方式二：SMTP
# ====================================================================

def build_email(mail_from, mail_to, mail_cc, subject, body, attachment_path=None):
    """构建 MIME 邮件对象（仅 SMTP 方式使用）。"""
    msg = MIMEMultipart()
    msg["From"] = mail_from
    msg["To"] = ", ".join(mail_to)
    if mail_cc:
        msg["Cc"] = ", ".join(mail_cc)
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "plain", "utf-8"))

    if attachment_path:
        file_path = Path(attachment_path)
        if not file_path.exists():
            raise FileNotFoundError(f"附件不存在: {attachment_path}")

        with open(file_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())

        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{file_path.name}"'
        )
        msg.attach(part)
        logger.info(f"已添加附件: {file_path.name} ({file_path.stat().st_size / 1024:.1f} KB)")

    return msg


def send_via_smtp(smtp_host, smtp_port, smtp_user, smtp_password,
                  msg, use_tls=True, max_retries=3):
    """通过 SMTP 协议发送邮件，支持重试。"""
    all_recipients = []
    if msg["To"]:
        all_recipients += [addr.strip() for addr in msg["To"].split(",")]
    if msg["Cc"]:
        all_recipients += [addr.strip() for addr in msg["Cc"].split(",")]

    for attempt in range(1, max_retries + 1):
        try:
            logger.info(f"正在通过 SMTP 发送邮件（第 {attempt} 次尝试）-> {all_recipients}")

            if smtp_port == 465:
                server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30)
            else:
                server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
                if use_tls:
                    server.starttls()

            if smtp_user and smtp_password:
                server.login(smtp_user, smtp_password)

            server.sendmail(msg["From"], all_recipients, msg.as_string())
            server.quit()

            logger.info("邮件通过 SMTP 发送成功")
            return True

        except smtplib.SMTPAuthenticationError:
            logger.error("SMTP 认证失败，请检查用户名和密码")
            raise

        except Exception as e:
            logger.warning(f"第 {attempt} 次发送失败: {e}")
            if attempt == max_retries:
                logger.error(f"邮件发送失败，已重试 {max_retries} 次")
                raise

    return False


# ====================================================================
# 内部工具：生成单个 GEO 的 Excel 附件
# ====================================================================

def _export_geo_excel(geo_df, geo_name, output_dir, ap_upload_sheets=None):
    """
    将单个 GEO 的推荐数据导出为 Excel 文件。
    包含主表的全部列供 GEO 参考，仅去掉 Geo 列。
    AP 额外包含各 SubGeo 的 BY Working 上传模板 sheet。
    """
    
    drop_cols = ["Geo"]
    if geo_name != "AP":
        drop_cols += ["SubGeo_Hist_Qty", "AP_Hist_Qty", "SubGeo_Share", "AP_Geo_Recommended_RR"]
    keep_cols = [c for c in geo_df.columns if c not in drop_cols]

    export_df = geo_df[keep_cols].copy()
    export_df = export_df.sort_values("PN").reset_index(drop=True)

    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"RR_Recommendation_{geo_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Recommendation", index=False)

        # AP: 附加各 SubGeo 的 BY Working 上传模板
        if geo_name == "AP" and ap_upload_sheets:
            for subgeo, sg_df in ap_upload_sheets.items():
                sheet_name = f"BY Working-{subgeo}"[:31]
                sg_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # ===== clour =====
        from openpyxl.styles import PatternFill
        ws_rec = writer.sheets["Recommendation"]
        
        grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        headers = [cell.value for cell in ws_rec[1]]
        
        for col_idx, header in enumerate(headers, 1):
            if header is None:
                continue
            h = str(header)
            if h.startswith("RR-"):
                fill = grey_fill
            elif h == "Recommended_RR" or (len(h) >= 5 and h[:4].isdigit() and h[4] == "-"):
                fill = green_fill
            else:
                continue
            for row in range(1, ws_rec.max_row + 1):
                ws_rec.cell(row=row, column=col_idx).fill = fill

    logger.info(f"已生成 {geo_name} 附件: {filename} ({len(export_df)} 行)")
    return filepath


# ====================================================================
# 发送单封邮件（内部使用）
# ====================================================================

def _send_one_email(mail_to, mail_cc, subject, body, attachment_path, config):
    """根据 config.MAIL_METHOD 发送单封邮件。"""
    method = getattr(config, "MAIL_METHOD", "outlook").lower()

    if method == "outlook":
        send_via_outlook(
            mail_to=mail_to,
            mail_cc=mail_cc,
            subject=subject,
            body=body,
            attachment_path=attachment_path,
        )

    elif method == "smtp":
        msg = build_email(
            mail_from=config.MAIL_FROM,
            mail_to=mail_to,
            mail_cc=mail_cc,
            subject=subject,
            body=body,
            attachment_path=attachment_path,
        )
        send_via_smtp(
            smtp_host=config.SMTP_HOST,
            smtp_port=config.SMTP_PORT,
            smtp_user=config.SMTP_USER,
            smtp_password=config.SMTP_PASSWORD,
            msg=msg,
            use_tls=config.SMTP_USE_TLS,
        )

    else:
        raise ValueError(f"不支持的 MAIL_METHOD: '{method}'")


# ====================================================================
# 统一入口：按 GEO 拆分并分别发送
# ====================================================================

def send_recommendation_email(result, config):
    """
    按 GEO 拆分推荐结果，每个 GEO 只收到自己的数据。
    GEO → 收件人映射配置在 config.GEO_RECIPIENTS 中。
    """
    rec_df = result["recommendation_df"]
    today_str = datetime.now().strftime("%Y-%m-%d")
    data_max_month = result["time_windows"]["data_max_month"].strftime("%Y-%m")

    geo_recipients = getattr(config, "GEO_RECIPIENTS", {})
    if not geo_recipients:
        logger.warning("GEO_RECIPIENTS 未配置，跳过邮件发送")
        return

    # 用临时目录存放各 GEO 的附件，发完自动清理
    tmp_dir = tempfile.mkdtemp(prefix="rr_geo_mail_")
    logger.info(f"临时附件目录: {tmp_dir}")

    sent_count = 0
    skipped_geos = []

    for geo, recipients in geo_recipients.items():
        # 检查是否启用
        if not recipients.get("enabled", True):
            logger.info(f"GEO [{geo}] 已关闭 (enabled=False)，跳过")
            skipped_geos.append(geo)
            continue

        mail_to = recipients.get("to", [])
        mail_cc = recipients.get("cc", [])

        if not mail_to:
            logger.warning(f"GEO [{geo}] 没有配置收件人，跳过")
            skipped_geos.append(geo)
            continue

        # 筛选该 GEO 的数据
        geo_df = rec_df.loc[rec_df["Geo"] == geo].copy()
        if geo_df.empty:
            logger.info(f"GEO [{geo}] 没有推荐数据，跳过")
            skipped_geos.append(geo)
            continue

        # 生成该 GEO 专属的 Excel 附件（包含该GEO的上传模板sheets，如果有）
        all_sheets = result.get("all_upload_sheets", {})
        geo_upload_sheets = all_sheets.get(geo) if geo in all_sheets else None
        attachment_path = _export_geo_excel(geo_df, geo, tmp_dir, ap_upload_sheets=geo_upload_sheets)

        subject = config.MAIL_SUBJECT_TEMPLATE.format(geo=geo, date=today_str, data_max_month=data_max_month)
        body = config.MAIL_BODY_TEMPLATE.format(
            geo=geo,
            date=today_str,
            data_max_month=data_max_month,
            recommendation_count=len(geo_df),
        )

        try:
            _send_one_email(mail_to, mail_cc, subject, body, attachment_path, config)
            sent_count += 1
        except Exception:
            logger.exception(f"GEO [{geo}] 邮件发送失败")

    # 清理临时文件
    try:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)
        logger.info("临时附件已清理")
    except Exception:
        logger.warning(f"临时目录清理失败: {tmp_dir}")

    logger.info(
        f"邮件发送完成: 成功 {sent_count} 个 GEO"
        + (f"，跳过 {skipped_geos}" if skipped_geos else "")
    )

    # 可选：发送汇总报告
    summary_to = getattr(config, "MAIL_TO_SUMMARY", [])
    summary_cc = getattr(config, "MAIL_CC_SUMMARY", [])
    if summary_to:
        subject = f"[Auto] L220 Low RR Recommendation - Full Report - {today_str}"
        body = (
            f"Hi,\n\n"
            f"The per-GEO recommendation emails have been sent.\n\n"
            f"Summary:\n"
            f"- Report Date: {today_str}\n"
            f"- Data Max Month: {data_max_month}\n"
            f"- Total Recommendation Rows: {len(rec_df)}\n"
            f"- GEOs Sent: {sent_count}\n"
            f"- GEOs Skipped: {skipped_geos if skipped_geos else 'None'}\n\n"
            f"Full report is attached.\n\n"
            f"Best Regards,\nChaozhong Xue"
        )
        try:
            _send_one_email(
                mail_to=summary_to,
                mail_cc=summary_cc,
                subject=subject,
                body=body,
                attachment_path=result["output_path"],
                config=config,
            )
            logger.info(f"汇总报告已发送 -> {summary_to}")
        except Exception:
            logger.exception("汇总报告发送失败")
