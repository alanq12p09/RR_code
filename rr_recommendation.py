"""
核心数据计算模块：RR数据提取、筛选、推荐值计算、结果导出。
变更：
  - EOL/LTB/LTS 不再排除，改为 Remark 列标记
  - NPI 从 KickOff 表识别（GA Date 在最近6个月内）
  - 取消条件3（6个月前有RR）
  - Recommendation 表增加历史月度RR列和未来3个月列
  - 有 Remark 的行推荐值和未来3个月显示 '-'
"""
import re
import logging
import pandas as pd
import numpy as np
from presto_conn import query_df

logger = logging.getLogger(__name__)


# ====================================================================
# 数据提取
# ====================================================================

def get_rr_data():
    sql = """
SELECT
    substr(rsd, 1, 7) AS RSD_Month,
    CASE
        WHEN geo IN ('EUROPE', 'META') THEN 'EMEA'
        ELSE geo
    END AS Geo,
    CASE 
        WHEN sub_geo = 'GTAP' THEN region_new
        WHEN sub_geo = 'HTK' THEN 'CAPK'
        WHEN sub_geo = 'ASEAN' THEN 'CAPK'
        WHEN sub_geo = 'LJP' THEN 'JAPAN'
        WHEN sub_geo IN ('IN','INDO') THEN 'INDIA'
        ELSE sub_geo
    END AS SubGeo,
    mtm AS PN,
    SUM(line_qty) AS Qty
FROM prd_sc_common_model.v_pcsd_order_detail_cml_brand_option
WHERE
    substr(rsd, 1, 7) >= DATE_FORMAT(DATE_ADD('month', -12, NOW()), '%Y-%m')
    AND substr(rsd, 1, 7) <= DATE_FORMAT(NOW(), '%Y-%m')
    AND substr(rsd, 1, 4) <> '9999'
    AND shipping_point = 'SC03'
    AND brand = 'Option'
    AND order_type IN ('Customer Order', 'Replenishment', 'Free of Charge', 'Internal Order')
    AND site_name = 'LSSC'
    AND status <> 'Order Cancelled'
    AND product_group IN ('TBG', 'LBG')
    AND (so_route <> 'CXLSSC' OR so_route IS NULL)
    AND (end_customer_name <> 'Lenovo Information Products(Shenzhen)Co.' OR end_customer_name IS NULL)
GROUP BY
    substr(rsd, 1, 7),
    CASE
        WHEN geo IN ('EUROPE', 'META') THEN 'EMEA'
        ELSE geo
    END,
    CASE 
        WHEN sub_geo = 'GTAP' THEN region_new
        WHEN sub_geo = 'HTK' THEN 'CAPK'
        WHEN sub_geo = 'ASEAN' THEN 'CAPK'
        WHEN sub_geo = 'LJP' THEN 'JAPAN'
        WHEN sub_geo IN ('IN','INDO') THEN 'INDIA'
        ELSE sub_geo
    END,
    mtm
    """
    return query_df(sql)


# ====================================================================
# 数据清洗
# ====================================================================

def load_pn_info(pn_info_path):
    df = pd.read_excel(pn_info_path)
    df.columns = [str(c).strip() for c in df.columns]

    rename_map = {}
    for c in df.columns:
        cu = c.upper()
        if cu in ["PN", "MTM", "PART NUMBER", "MATERIAL"]:
            rename_map[c] = "PN"
        elif cu in ["COMMENTS", "COMMENT", "REMARK", "REMARKS"]:
            rename_map[c] = "Comments"
        elif cu in ["CATEGORY1", "CATEGORY 1", "BU"]:
            rename_map[c] = "Category1"

    df = df.rename(columns=rename_map)

    required = ["PN", "Comments", "Category1"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"PN_INFO 缺少必要列: {missing}")

    df["PN"] = df["PN"].astype(str).str.strip()
    df["Comments"] = df["Comments"].fillna("").astype(str)
    df["Category1"] = df["Category1"].fillna("").astype(str).str.upper().str.strip()
    return df


def load_kickoff(kickoff_path):
    """读取 KickOff Package 表，提取 PN 和 GA Date。"""
    df = pd.read_excel(kickoff_path)
    df.columns = [str(c).strip() for c in df.columns]

    # 查找 PN 列
    pn_col = None
    for c in df.columns:
        if "orderable" in c.lower() and "pn" in c.lower():
            pn_col = c
            break
    if pn_col is None:
        # 尝试其他可能的列名
        for c in df.columns:
            if c.upper() in ["PN", "MTM", "PART NUMBER"]:
                pn_col = c
                break
    if pn_col is None:
        raise ValueError(f"KickOff 表找不到 PN 列，现有列: {list(df.columns)}")

    # 查找 GA Date 列
    ga_col = None
    for c in df.columns:
        cl = c.lower()
        if "launch" in cl or "ga" in cl:
            ga_col = c
            break
    if ga_col is None:
        raise ValueError(f"KickOff 表找不到 GA Date 列，现有列: {list(df.columns)}")

    result = df[[pn_col, ga_col]].copy()
    result.columns = ["PN", "GA_Date"]
    result["PN"] = result["PN"].astype(str).str.strip()
    result["GA_Date"] = pd.to_datetime(result["GA_Date"], errors="coerce")
    result = result.dropna(subset=["GA_Date"])

    logger.info(f"KickOff 表读取完成: {len(result)} 行，PN列='{pn_col}'，GA列='{ga_col}'")
    return result


def prepare_rr(rr_raw):
    rr = rr_raw.copy()
    rr.columns = [str(c).strip() for c in rr.columns]

    required = ["RSD_Month", "Geo", "SubGeo", "PN", "Qty"]
    missing = [c for c in required if c not in rr.columns]
    if missing:
        raise ValueError(f"RR 数据缺少必要列: {missing}")

    rr["RSD_Month"] = pd.to_datetime(rr["RSD_Month"].astype(str).str[:7] + "-01", errors="coerce")
    rr["Geo"] = rr["Geo"].astype(str).str.strip()
    rr["SubGeo"] = rr["SubGeo"].astype(str).str.strip()
    rr["PN"] = rr["PN"].astype(str).str.strip()
    rr["Qty"] = pd.to_numeric(rr["Qty"], errors="coerce").fillna(0)

    rr = rr.dropna(subset=["RSD_Month"])

    # 双重保障：截断超过当前月份的数据
    current_month = pd.Timestamp.now().to_period("M").to_timestamp()
    before_count = len(rr)
    rr = rr.loc[rr["RSD_Month"] <= current_month].copy()
    dropped = before_count - len(rr)
    if dropped > 0:
        logger.warning(f"已截断 {dropped} 行未来月份数据")

    return rr


def filter_pn_info(pn_info):
    """
    不再排除 EOL/LTB/LTS，只排除 DOCK 品类。
    EOL/LTB/LTS 状态在后续通过 Remark 列标记。
    """
    df = pn_info.copy()
    df["Exclude_By_DOCK"] = df["Category1"].eq("DOCK")
    df["Valid_PN"] = ~df["Exclude_By_DOCK"]

    valid_pn = df.loc[df["Valid_PN"], ["PN"]].drop_duplicates()
    return df, valid_pn


def build_remark(pn_info, kickoff_df, current_month, rr_filtered=None):
    pattern = re.compile(r"(EOL|LTB|LTS)", flags=re.IGNORECASE)
    pn_remarks = {}

    for _, row in pn_info.iterrows():
        pn = str(row["PN"]).strip()
        comments = str(row.get("Comments", ""))
        if pattern.search(comments):
            pn_remarks[pn] = "EOL/LTB"

    # NPI来源1: KickOff表 GA Date 在最近6个月内
    cutoff_6m_ago = (current_month.to_period("M") - 6).to_timestamp()
    npi_pns = kickoff_df.loc[
        kickoff_df["GA_Date"] >= cutoff_6m_ago, "PN"
    ].unique()

    for pn in npi_pns:
        if pn not in pn_remarks:
            pn_remarks[pn] = "NPI"

    # NPI来源2: 6个月前及更早没有RR记录的PN
    if rr_filtered is not None:
        pn_first_month = rr_filtered.groupby("PN")["RSD_Month"].min()
        for pn, first_month in pn_first_month.items():
            if first_month > cutoff_6m_ago and pn not in pn_remarks:
                pn_remarks[pn] = "NPI"

    return pn_remarks


# ====================================================================
# 时间窗口
# ====================================================================

def get_time_windows(rr):
    current_month = pd.Timestamp.now().to_period("M").to_timestamp()
    current_quarter = current_month.to_period("Q")

    data_max_month = rr["RSD_Month"].max()
    if data_max_month < (current_month.to_period("M") - 1).to_timestamp():
        logger.warning(
            f"数据最新月份为 {data_max_month:%Y-%m}，"
            f"落后于系统当前月份 {current_month:%Y-%m} 超过1个月"
        )

    recent_4m = [(current_month.to_period("M") - i).to_timestamp() for i in range(4)]
    recent_4m = sorted(recent_4m)

    prev_4_quarters = [current_quarter - i for i in range(1, 5)]

    # 未来3个月（含本月）
    future_3m = [(current_month.to_period("M") + i).to_timestamp() for i in range(4)]

    return {
        "current_month": current_month,
        "current_quarter": current_quarter,
        "recent_4m": recent_4m,
        "prev_4_quarters": prev_4_quarters,
        "future_3m": future_3m,
        "data_max_month": data_max_month,
    }


# ====================================================================
# 核心筛选（条件3已取消）
# ====================================================================

def build_geo_pn_summary(rr, valid_pn, time_windows, min_recent_months=3):
    rr2 = rr.merge(valid_pn, on="PN", how="inner").copy()
    rr2["Quarter"] = rr2["RSD_Month"].dt.to_period("Q")

    recent_4m = set(time_windows["recent_4m"])
    prev_4_quarters = time_windows["prev_4_quarters"]

    geo_pn_month = rr2.groupby(["PN", "Geo", "RSD_Month"], as_index=False)["Qty"].sum()
    geo_pn_qtr = rr2.groupby(["PN", "Geo", "Quarter"], as_index=False)["Qty"].sum()

    pn_total = (
        rr2.groupby("PN", as_index=False)["Qty"].sum()
        .rename(columns={"Qty": "PN_Total_Qty_AllGeo"})
    )
    geo_total = (
        rr2.groupby(["PN", "Geo"], as_index=False)["Qty"].sum()
        .rename(columns={"Qty": "PN_Geo_Total_Qty"})
    )

    summary = geo_total.merge(pn_total, on="PN", how="left")
    summary["Geo_Share_of_PN"] = np.where(
        summary["PN_Total_Qty_AllGeo"] > 0,
        summary["PN_Geo_Total_Qty"] / summary["PN_Total_Qty_AllGeo"],
        0
    )
    '''''
    # 条件1：过去4个季度（不含当前季度）平均RR < 500
    for q in prev_4_quarters:
        temp = geo_pn_qtr.loc[geo_pn_qtr["Quarter"] == q, ["PN", "Geo", "Qty"]].copy()
        temp.rename(columns={"Qty": f"Qty_{str(q)}"}, inplace=True)
        summary = summary.merge(temp, on=["PN", "Geo"], how="left")

    q_cols = [c for c in summary.columns if c.startswith("Qty_20") and "Q" in c]
    for c in q_cols:
        summary[c] = summary[c].fillna(0)

    #summary["Avg_Last4Q_Excl_CQ"] = summary[q_cols].mean(axis=1) if q_cols else 0
    #summary["Flag_Avg_Last4Q_LT_500"] = summary["Avg_Last4Q_Excl_CQ"] < 500
    if q_cols:
        q_data = summary[q_cols]
        q_nonzero_count = (q_data > 0).sum(axis=1).replace(0, np.nan)
        summary["Avg_Last4Q_Excl_CQ"] = q_data.sum(axis=1) / q_nonzero_count
    else:
        summary["Avg_Last4Q_Excl_CQ"] = 0
    summary["Flag_Avg_Last4Q_LT_500"] = summary["Avg_Last4Q_Excl_CQ"] < 500

     '''''
    
    # 条件1：过去12个月（不含当前月），按3个月滚动切分为4段
    # 平均值 = 总量 / 有数据的段数，< 500 为低量
    current_month = time_windows["current_month"]
    current_period = current_month.to_period("M")

    # 4个滚动季度：[-12,-10], [-9,-7], [-6,-4], [-3,-1]（相对当前月）
    rolling_quarters = []
    for i in range(4):
        start = (current_period - 12 + i * 3).to_timestamp()
        end = (current_period - 12 + (i + 1) * 3 - 1).to_timestamp()  # 包含该月
        rolling_quarters.append((start, end))

    # 每段的合计
    for idx, (q_start, q_end) in enumerate(rolling_quarters):
        q_label = f"RollingQ{idx+1}"
        temp = (
            geo_pn_month.loc[
                (geo_pn_month["RSD_Month"] >= q_start) &
                (geo_pn_month["RSD_Month"] <= q_end)
            ]
            .groupby(["PN", "Geo"], as_index=False)["Qty"].sum()
            .rename(columns={"Qty": q_label})
        )
        summary = summary.merge(temp, on=["PN", "Geo"], how="left")

    rq_cols = [f"RollingQ{i+1}" for i in range(4)]
    for c in rq_cols:
        summary[c] = summary[c].fillna(0)

    rq_data = summary[rq_cols]
    rq_nonzero_count = (rq_data > 0).sum(axis=1).replace(0, np.nan)
    summary["Avg_RollingQ"] = rq_data.sum(axis=1) / rq_nonzero_count
    summary["Avg_RollingQ"] = summary["Avg_RollingQ"].fillna(0)
    summary["RollingQ_With_Data"] = rq_nonzero_count.fillna(0).astype(int)
    summary["Flag_Avg_RollingQ_LT_500"] = summary["Avg_RollingQ"] < 500

    # 条件2：最近4个月中至少 min_recent_months 个月有RR
    recent_rr_detail = geo_pn_month.loc[geo_pn_month["RSD_Month"].isin(recent_4m)].copy()
    recent_rr_check = (
        recent_rr_detail.groupby(["PN", "Geo"], as_index=False)
        .agg(
            Qty_Recent4M=("Qty", "sum"),
            Recent4M_Month_Count=("RSD_Month", "nunique")
        )
    )

    summary = summary.merge(recent_rr_check, on=["PN", "Geo"], how="left")
    summary["Qty_Recent4M"] = summary["Qty_Recent4M"].fillna(0)
    summary["Recent4M_Month_Count"] = summary["Recent4M_Month_Count"].fillna(0).astype(int)
    summary["Flag_Recent4M_HasRR"] = summary["Recent4M_Month_Count"] >= min_recent_months


    # 条件3：历史有效RR月份数 >= 4（数据太少的排除）
    months_with_rr = (
        geo_pn_month.groupby(["PN", "Geo"], as_index=False)
        .agg(Total_Months_With_RR=("RSD_Month", "nunique"))
    )
    summary = summary.merge(months_with_rr, on=["PN", "Geo"], how="left")
    summary["Total_Months_With_RR"] = summary["Total_Months_With_RR"].fillna(0).astype(int)
    summary["Flag_Enough_History"] = summary["Total_Months_With_RR"] >= 4

    # 条件4：PN总量>500，但某GEO占比<3%，排除该GEO
    summary["Exclude_Geo_LowShare_When_PNHigh"] = (
        (summary["PN_Total_Qty_AllGeo"] > 500) &
        (summary["Geo_Share_of_PN"] < 0.03)
    )

    summary["Qualified"] = (
        summary["Flag_Avg_RollingQ_LT_500"] &
        summary["Flag_Recent4M_HasRR"] &
        summary["Flag_Enough_History"] &
        (~summary["Exclude_Geo_LowShare_When_PNHigh"])
    )

    return summary, rr2


# ====================================================================
# 推荐值计算
# ====================================================================

def trimmed_mean(series):
    vals = pd.to_numeric(series, errors="coerce").dropna().tolist()
    vals = sorted(vals)
    if len(vals) == 0:
        return np.nan
    if len(vals) == 1:
        return vals[0]
    if len(vals) == 2:
        return np.mean(vals)
    return np.mean(vals[1:-1])


def _get_recommendation_base(rr_filtered, summary, current_month, include_current_month=False):
    qualified_keys = summary.loc[summary["Qualified"], ["PN", "Geo"]].drop_duplicates()
    base = rr_filtered.merge(qualified_keys, on=["PN", "Geo"], how="inner").copy()
    if not include_current_month:
        base = base.loc[base["RSD_Month"] < current_month].copy()
    return base


def build_recommendation(rr_filtered, summary, time_windows, pn_remarks,
                         include_current_month=False,
                         ap_subgeo_min_rr=5,
                         ap_subgeo_min_share=0.03):
    """
    非AP：按 PN + GEO 直接算 trimmed_mean
    AP：先算 PN + AP，再按 SubGeo 历史占比分配
    增加：历史月度RR列、未来3个月列、Remark列
    """
    current_month = time_windows["current_month"]
    future_3m = time_windows["future_3m"]

    base = _get_recommendation_base(
        rr_filtered, summary, current_month, include_current_month
    )

    # ---------- 非 AP ----------
    non_ap = base.loc[base["Geo"] != "AP"].copy()
    non_ap_month = non_ap.groupby(["PN", "Geo", "RSD_Month"], as_index=False)["Qty"].sum()

    # 补零计算：从首次出现到上个月，没有RR的月份补0参与trimmed_mean
    all_months = pd.date_range(
        non_ap_month["RSD_Month"].min(),
        current_month - pd.DateOffset(months=1),
        freq="MS"
    )
    first_month_non_ap = non_ap_month.groupby(["PN", "Geo"])["RSD_Month"].min().reset_index(name="First_Month")
    full_non_ap = (
        non_ap_month[["PN", "Geo"]].drop_duplicates()
        .merge(pd.DataFrame({"RSD_Month": all_months}), how="cross")
        .merge(first_month_non_ap, on=["PN", "Geo"])
    )
    full_non_ap = full_non_ap.loc[full_non_ap["RSD_Month"] >= full_non_ap["First_Month"]].drop(columns="First_Month")
    non_ap_month_full = full_non_ap.merge(non_ap_month, on=["PN", "Geo", "RSD_Month"], how="left")
    non_ap_month_full["Qty"] = non_ap_month_full["Qty"].fillna(0)

    non_ap_rec = (
        non_ap_month_full.groupby(["PN", "Geo"])["Qty"]
        .apply(trimmed_mean)
        .reset_index(name="Recommended_RR")
    )

    non_ap_rec["SubGeo"] = ""
    non_ap_rec["Recommend_Level"] = "GEO"
    non_ap_rec["Recommend_To"] = non_ap_rec["Geo"]

    non_ap_stats = (
        non_ap_month_full.groupby(["PN", "Geo"])["Qty"]
        .agg(Month_Count="count", Hist_Min="min", Hist_Max="max", Hist_Avg="mean")
        .reset_index()
    )
    non_ap_stats["SubGeo"] = ""

    non_ap_result = non_ap_rec.merge(
        non_ap_stats[["PN", "Geo", "SubGeo", "Month_Count", "Hist_Min", "Hist_Max", "Hist_Avg"]],
        on=["PN", "Geo", "SubGeo"], how="left"
    )

    # 非AP历史月度pivot
    non_ap_hist_pivot = non_ap_month.pivot_table(
        index=["PN", "Geo"], columns="RSD_Month", values="Qty", aggfunc="sum"
    ).reset_index()
    non_ap_hist_pivot.columns.name = None
    non_ap_hist_pivot["SubGeo"] = ""
    non_ap_result = non_ap_result.merge(non_ap_hist_pivot, on=["PN", "Geo", "SubGeo"], how="left")

    # ---------- AP ----------
    ap = base.loc[base["Geo"] == "AP"].copy()

    ap_month_total = ap.groupby(["PN", "Geo", "RSD_Month"], as_index=False)["Qty"].sum()

    # 补零计算
    if not ap_month_total.empty:
        all_months_ap = pd.date_range(
            ap_month_total["RSD_Month"].min(),
            current_month - pd.DateOffset(months=1),
            freq="MS"
        )
        first_month_ap = ap_month_total.groupby(["PN", "Geo"])["RSD_Month"].min().reset_index(name="First_Month")
        full_ap = (
            ap_month_total[["PN", "Geo"]].drop_duplicates()
            .merge(pd.DataFrame({"RSD_Month": all_months_ap}), how="cross")
            .merge(first_month_ap, on=["PN", "Geo"])
        )
        full_ap = full_ap.loc[full_ap["RSD_Month"] >= full_ap["First_Month"]].drop(columns="First_Month")
        ap_month_total_full = full_ap.merge(ap_month_total, on=["PN", "Geo", "RSD_Month"], how="left")
        ap_month_total_full["Qty"] = ap_month_total_full["Qty"].fillna(0)
    else:
        ap_month_total_full = ap_month_total

    ap_geo_rec = (
        ap_month_total_full.groupby(["PN", "Geo"])["Qty"]
        .apply(trimmed_mean)
        .reset_index(name="AP_Geo_Recommended_RR")
    )

    ap_geo_stats = (
        ap_month_total_full.groupby(["PN", "Geo"])["Qty"]
        .agg(AP_Month_Count="count", AP_Hist_Min="min", AP_Hist_Max="max", AP_Hist_Avg="mean")
        .reset_index()
    )

    ap_subgeo_hist = (
        ap.groupby(["PN", "Geo", "SubGeo"], as_index=False)["Qty"].sum()
        .rename(columns={"Qty": "SubGeo_Hist_Qty"})
    )
    ap_pn_hist = (
        ap.groupby(["PN", "Geo"], as_index=False)["Qty"].sum()
        .rename(columns={"Qty": "AP_Hist_Qty"})
    )

    ap_alloc = ap_subgeo_hist.merge(ap_pn_hist, on=["PN", "Geo"], how="left")
    ap_alloc["SubGeo_Share"] = np.where(
        ap_alloc["AP_Hist_Qty"] > 0,
        ap_alloc["SubGeo_Hist_Qty"] / ap_alloc["AP_Hist_Qty"], 0
    )

    ap_result = (
        ap_alloc
        .merge(ap_geo_rec, on=["PN", "Geo"], how="left")
        .merge(ap_geo_stats, on=["PN", "Geo"], how="left")
    )

    ap_result["Recommended_RR"] = ap_result["AP_Geo_Recommended_RR"] * ap_result["SubGeo_Share"]
    ap_result["Recommend_Level"] = "SubGeo_Allocated_From_AP"
    ap_result["Recommend_To"] = ap_result["SubGeo"]
    ap_result["Month_Count"] = ap_result["AP_Month_Count"]
    ap_result["Hist_Min"] = ap_result["AP_Hist_Min"]
    ap_result["Hist_Max"] = ap_result["AP_Hist_Max"]
    ap_result["Hist_Avg"] = ap_result["AP_Hist_Avg"]

    # 双重过滤
    ap_before = len(ap_result)
    ap_result = ap_result.loc[
        (ap_result["Recommended_RR"] >= ap_subgeo_min_rr) &
        (ap_result["SubGeo_Share"] >= ap_subgeo_min_share)
    ].copy()
    ap_dropped = ap_before - len(ap_result)
    if ap_dropped > 0:
        logger.info(f"AP SubGeo 双重过滤: {ap_dropped} 行被移除")

    # AP SubGeo级历史月度pivot
    ap_subgeo_month = ap.groupby(["PN", "Geo", "SubGeo", "RSD_Month"], as_index=False)["Qty"].sum()
    ap_hist_pivot = ap_subgeo_month.pivot_table(
        index=["PN", "Geo", "SubGeo"], columns="RSD_Month", values="Qty", aggfunc="sum"
    ).reset_index()
    ap_hist_pivot.columns.name = None
    ap_result = ap_result.merge(ap_hist_pivot, on=["PN", "Geo", "SubGeo"], how="left")

    # ---------- 合并 ----------
    non_ap_result["SubGeo_Hist_Qty"] = np.nan
    non_ap_result["AP_Hist_Qty"] = np.nan
    non_ap_result["SubGeo_Share"] = np.nan
    non_ap_result["AP_Geo_Recommended_RR"] = np.nan

    rec = pd.concat([non_ap_result, ap_result], ignore_index=True)

    # ---------- Remark 列 ----------
    rec["Remark"] = rec["PN"].map(pn_remarks).fillna("NULL")

    # ---------- 未来3个月列 ----------
    for fm in future_3m:
        col_name = fm.strftime("%Y-%b")
        rec[col_name] = np.where(
            rec["Remark"] == "NULL",
            rec["Recommended_RR"],
            np.nan  # 先用nan占位，后面统一替换
        )

    # ---------- 有 Remark 的行，推荐值和未来月显示 '-' ----------
    has_remark = rec["Remark"] != "NULL"
    rec.loc[has_remark, "Recommended_RR"] = np.nan

    # ---------- 整理历史月度列名：RSD_Month timestamp -> "RR-YYYY-MM" ----------
    month_cols_map = {}
    for c in rec.columns:
        if isinstance(c, pd.Timestamp):
            month_cols_map[c] = f"RR-{c.strftime('%Y-%m')}"
    rec = rec.rename(columns=month_cols_map)

    rr_cols = sorted([c for c in rec.columns if c.startswith("RR-")])
    future_cols = [fm.strftime("%Y-%b") for fm in future_3m]

    # ---------- 数值四舍五入 ----------
    num_cols_to_round = (
        ["Recommended_RR", "Hist_Avg", "Hist_Min", "Hist_Max", "AP_Geo_Recommended_RR"]
        + rr_cols + future_cols
    )
    for c in num_cols_to_round:
        if c in rec.columns:
            rec[c] = pd.to_numeric(rec[c], errors="coerce")
            rec[c] = rec[c].round(0)

    if "SubGeo_Share" in rec.columns:
        rec["SubGeo_Share"] = pd.to_numeric(rec["SubGeo_Share"], errors="coerce").round(4)

    # ---------- 有 Remark 的行数值列替换为 '-' ----------
    display_cols = ["Recommended_RR"] + future_cols
    for c in display_cols:
        if c in rec.columns:
            rec[c] = rec[c].astype(object)
            rec.loc[has_remark, c] = "-"

    # ---------- 整理列顺序 ----------
    base_cols = ["PN", "Geo", "SubGeo", "Recommend_Level", "Recommend_To"]
    mid_cols = rr_cols + ["Recommended_RR"] + future_cols + ["Remark"]
    stat_cols = ["Month_Count", "Hist_Min", "Hist_Max", "Hist_Avg",
                 "SubGeo_Hist_Qty", "AP_Hist_Qty", "SubGeo_Share", "AP_Geo_Recommended_RR"]

    final_cols = base_cols + mid_cols + stat_cols
    # 只保留实际存在的列
    final_cols = [c for c in final_cols if c in rec.columns]

    rec = rec[final_cols].sort_values(["PN", "Geo", "SubGeo"]).reset_index(drop=True)

    return rec


# ====================================================================
# 历史参考
# ====================================================================

def build_history_reference(rr_filtered, summary):
    qualified_keys = summary.loc[summary["Qualified"], ["PN", "Geo"]].drop_duplicates()
    hist = rr_filtered.merge(qualified_keys, on=["PN", "Geo"], how="inner")
    hist = (
        hist.groupby(["PN", "Geo", "SubGeo", "RSD_Month"], as_index=False)["Qty"]
        .sum()
        .sort_values(["PN", "Geo", "SubGeo", "RSD_Month"])
    )
    hist["RSD_Month"] = hist["RSD_Month"].dt.strftime("%Y-%m")
    return hist


# ====================================================================
# GEO 上传模板构建器（通用框架，按 GEO 注册不同模板）
# ====================================================================
#
# 扩展方式：当其他 GEO 需要上传模板时：
#   1. 在下面新建一个 build_XX_upload_sheets() 函数（参考 build_ap_upload_sheets）
#   2. 在 GEO_UPLOAD_BUILDERS 字典里注册：GEO_UPLOAD_BUILDERS["XX"] = build_XX_upload_sheets
#   3. 在 config.py 里加该 GEO 的模板配置（source、overrides 等）
#   4. run_recommendation 会自动调用
#
# 例如 EMEA 需要不同格式的模板：
#   def build_emea_upload_sheets(recommendation, time_windows, **kwargs):
#       ... 按 EMEA 模板格式生成 ...
#       return {"sheet_name": DataFrame}
#   GEO_UPLOAD_BUILDERS["EMEA"] = build_emea_upload_sheets
# ====================================================================

def build_ap_upload_sheets(recommendation, time_windows,
                           source="0001000835", overrides=None, **kwargs):
    """
    AP 上传模板：按 SubGeo 分 sheet，套用 BY Working 格式。
    排除有 Remark 的 PN，前3个月填推荐值，后6个月填 '-'。
    """
    if overrides is None:
        overrides = {}

    ap_rec = recommendation.loc[recommendation["Geo"] == "AP"].copy()
    if ap_rec.empty:
        return {}

    # 排除有 Remark 的 PN（EOL/LTB/LTS/NPI 不出现在上传表中）
    if "Remark" in ap_rec.columns:
        before = len(ap_rec)
        ap_rec = ap_rec.loc[ap_rec["Remark"] == "NULL"].copy()
        excluded = before - len(ap_rec)
        if excluded > 0:
            logger.info(f"AP 上传模板: 排除 {excluded} 行有 Remark 的 PN")

    if ap_rec.empty:
        return {}

    # 未来9个月的列名
    current_period = time_windows["current_month"].to_period("M")
    future_months = [(current_period + i).to_timestamp() for i in range(9)]
    future_col_names = [m.strftime("%Y-%b") for m in future_months]

    subgeo_sheets = {}

    for subgeo in sorted(ap_rec["SubGeo"].unique()):
        if not subgeo or subgeo.strip() == "":
            continue

        sg_df = ap_rec.loc[ap_rec["SubGeo"] == subgeo].copy()

        rows = []
        for _, row in sg_df.iterrows():
            pn = row["PN"]

            # 覆盖值优先
            override_val = overrides.get(pn, {}).get(subgeo, None)

            # 推荐值
            rec_val = row.get("Recommended_RR", None)

            month_vals = []
            for i in range(9):
                if i < 4:
                    # 前3个月：覆盖值 > 推荐值
                    if override_val is not None:
                        month_vals.append(int(round(override_val)))
                    elif rec_val is not None and rec_val != "-" and pd.notna(rec_val):
                        month_vals.append(int(round(float(rec_val))))
                    else:
                        month_vals.append("")
                else:
                    # 后6个月填 '-'
                    month_vals.append("-")

            r = {
                "Item Code": pn,
                "Country Group": subgeo,
                "Customer Segment": None,
                "Channel": None,
                "Biz Unit": None,
                "Source": source,
                "Comments": None,
            }
            for col_name, val in zip(future_col_names, month_vals):
                r[col_name] = val

            rows.append(r)

        sheet_df = pd.DataFrame(rows)
        sheet_df = sheet_df.sort_values("Item Code").reset_index(drop=True)
        subgeo_sheets[subgeo] = sheet_df

    return subgeo_sheets


# 注册表：GEO -> 模板构建函数
# 新增其他 GEO 模板时，在这里注册即可
GEO_UPLOAD_BUILDERS = {
    "AP": build_ap_upload_sheets,
    # "EMEA": build_emea_upload_sheets,   # 示例：以后 EMEA 需要时取消注释
    # "NA":   build_na_upload_sheets,     # 示例：以后 NA 需要时取消注释
}


def build_all_upload_sheets(recommendation, time_windows, upload_configs=None):
    """
    统一入口：遍历 GEO_UPLOAD_BUILDERS，为已注册的 GEO 生成上传模板。

    Args:
        recommendation: 完整推荐表
        time_windows: 时间窗口
        upload_configs: { "GEO": { 该GEO模板所需的参数 } }
            例如 { "AP": { "source": "0001000835", "overrides": {} } }

    Returns:
        dict: { "GEO": { "sheet_name": DataFrame } }
    """
    if upload_configs is None:
        upload_configs = {}

    all_sheets = {}
    for geo, builder_fn in GEO_UPLOAD_BUILDERS.items():
        geo_config = upload_configs.get(geo, {})
        try:
            sheets = builder_fn(recommendation, time_windows, **geo_config)
            if sheets:
                all_sheets[geo] = sheets
                logger.info(f"{geo} 上传模板: {len(sheets)} 个 sheet")
        except Exception:
            logger.exception(f"{geo} 上传模板生成失败")

    return all_sheets


# ====================================================================
# 导出
# ====================================================================

def export_results(output_path, summary, history, recommendation, config_df,
                   all_upload_sheets=None):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        config_df.to_excel(writer, sheet_name="Config", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        summary.loc[summary["Qualified"]].to_excel(writer, sheet_name="Qualified", index=False)
        history.to_excel(writer, sheet_name="History_RR", index=False)
        recommendation.to_excel(writer, sheet_name="Recommendation", index=False)

        # 各 GEO 上传模板 sheets
        if all_upload_sheets:
            for geo, sheets in all_upload_sheets.items():
                for sheet_label, df in sheets.items():
                    sheet_name = f"BY Working-{sheet_label}"[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 给 Recommendation sheet 上色
        from openpyxl.styles import PatternFill
        ws_rec = writer.sheets["Recommendation"]
        
        grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # 读取表头，找到对应列的位置
        headers = [cell.value for cell in ws_rec[1]]
        
        for col_idx, header in enumerate(headers, 1):
            if header is None:
                continue
            h = str(header)
            if h.startswith("RR-"):
                fill = grey_fill
            elif h in ("Recommended_RR",) or any(
                h.startswith(f"{y}-") for y in range(2025, 2035)
            ):
                fill = green_fill
            else:
                continue
            # 从表头到最后一行都涂色
            for row in range(1, ws_rec.max_row + 1):
                ws_rec.cell(row=row, column=col_idx).fill = fill

    logger.info(f"结果已保存到: {output_path}")


# ====================================================================
# 主流程
# ====================================================================

def run_recommendation(pn_info_path, output_path, kickoff_path,
                       include_current_month=False,
                       min_recent_months=3,
                       ap_subgeo_min_rr=5,
                       ap_subgeo_min_share=0.03,
                       upload_configs=None):
    """执行完整的推荐值计算流程。"""
    if upload_configs is None:
        upload_configs = {}

    rr_raw = get_rr_data()
    rr = prepare_rr(rr_raw)

    pn_info = load_pn_info(pn_info_path)
    _, valid_pn = filter_pn_info(pn_info)

    kickoff_df = load_kickoff(kickoff_path)

    time_windows = get_time_windows(rr)

    summary, rr_filtered = build_geo_pn_summary(
        rr, valid_pn, time_windows,
        min_recent_months=min_recent_months
    )

    pn_remarks = build_remark(pn_info, kickoff_df, time_windows["current_month"],
                              rr_filtered=rr_filtered)
    logger.info(
        f"Remark 标记: EOL/LTB={sum(1 for v in pn_remarks.values() if v == 'EOL/LTB')}, "
        f"NPI={sum(1 for v in pn_remarks.values() if v == 'NPI')}"
    )

    history = build_history_reference(rr_filtered, summary)

    recommendation = build_recommendation(
        rr_filtered=rr_filtered,
        summary=summary,
        time_windows=time_windows,
        pn_remarks=pn_remarks,
        include_current_month=include_current_month,
        ap_subgeo_min_rr=ap_subgeo_min_rr,
        ap_subgeo_min_share=ap_subgeo_min_share
    )

    # 生成各 GEO 上传模板（通用框架）
    all_upload_sheets = build_all_upload_sheets(
        recommendation=recommendation,
        time_windows=time_windows,
        upload_configs=upload_configs,
    )

    config_df = pd.DataFrame({
        "Parameter": [
            "INCLUDE_CURRENT_MONTH_IN_RECOMMENDATION",
            "MIN_RECENT_MONTHS_WITH_RR",
            "AP_SUBGEO_MIN_RECOMMENDED_RR",
            "AP_SUBGEO_MIN_SHARE",
            "Current_Month (System Time)",
            "Data_Max_Month",
            "Recent_4M_Rule",
            "AP_Recommendation_Method",
            "Remark_Rule",
        ],
        "Value": [
            include_current_month,
            min_recent_months,
            ap_subgeo_min_rr,
            ap_subgeo_min_share,
            time_windows["current_month"].strftime("%Y-%m"),
            time_windows["data_max_month"].strftime("%Y-%m"),
            f"最近4个月中至少 {min_recent_months} 个月有RR",
            f"先算AP的Geo recommendation，再按SubGeo历史占比分配，"
            f"双重过滤：绝对值 < {ap_subgeo_min_rr} 或占比 < {ap_subgeo_min_share:.0%} 的行被移除",
            "EOL/LTB/LTS从Comments识别，NPI从KickOff表GA Date最近6个月识别，有Remark的不提供建议值",
        ]
    })

    export_results(output_path, summary, history, recommendation, config_df,
                   all_upload_sheets=all_upload_sheets)

    result = {
        "output_path": output_path,
        "time_windows": time_windows,
        "qualified_count": int(summary["Qualified"].sum()),
        "recommendation_count": len(recommendation),
        "recommendation_df": recommendation,
        "all_upload_sheets": all_upload_sheets,
        "include_current_month": include_current_month,
        "min_recent_months": min_recent_months,
        "ap_subgeo_min_rr": ap_subgeo_min_rr,
        "ap_subgeo_min_share": ap_subgeo_min_share,
    }

    logger.info(
        f"计算完成 | Qualified: {result['qualified_count']:,} | "
        f"Recommendations: {result['recommendation_count']:,} | "
        f"Upload GEOs: {list(all_upload_sheets.keys())}"
    )

    return result


# ====================================================================
# 单独运行
# ====================================================================

if __name__ == "__main__":
    import config

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s - %(message)s"
    )

    result = run_recommendation(
        pn_info_path=config.PN_INFO_PATH,
        output_path=config.OUTPUT_PATH,
        kickoff_path=config.KICKOFF_PATH,
        include_current_month=config.INCLUDE_CURRENT_MONTH_IN_RECOMMENDATION,
        min_recent_months=config.MIN_RECENT_MONTHS_WITH_RR,
        ap_subgeo_min_rr=config.AP_SUBGEO_MIN_RECOMMENDED_RR,
        ap_subgeo_min_share=config.AP_SUBGEO_MIN_SHARE,
        upload_configs=config.UPLOAD_CONFIGS,
    )

    print("=" * 100)
    print(f"结果已保存到: {result['output_path']}")
    print(f"Current Month: {result['time_windows']['current_month'].strftime('%Y-%m')}")
    print(f"Qualified GEO+PN: {result['qualified_count']:,}")
    print(f"Recommendation Rows: {result['recommendation_count']:,}")
    print("=" * 100)
